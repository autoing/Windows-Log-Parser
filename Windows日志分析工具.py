from Evtx import Evtx
import time
import openpyxl
from xml.etree import ElementTree
from alive_progress import alive_bar
from datetime import datetime,timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

filename = '日志分析结果.xlsx'
wb = openpyxl.Workbook()
# 登录日志
wslogin = wb.active
wslogin.freeze_panes = 'A2'
wslogin.title = '登录日志'
wslogin.column_dimensions['A'].width = 20
wslogin.column_dimensions['B'].width = 12
wslogin.column_dimensions['C'].width = 12
wslogin.column_dimensions['D'].width = 20
wslogin.column_dimensions['E'].width = 20
wslogin.column_dimensions['F'].width = 12
wslogin.column_dimensions['G'].width = 50
wslogin.row_dimensions[1].height = 20
wslogin['A1'] = '时间'
wslogin['B1'] = '事件ID'
wslogin['C1'] = '登录状态'
wslogin['D1'] = '用户名'
wslogin['E1'] = '登录IP'
wslogin['F1'] = '登录类型'
wslogin['G1'] = '进程名称'

#账号管理
wb.create_sheet('账号管理')
wsuser = wb.get_sheet_by_name('账号管理')
wsuser.column_dimensions['A'].width = 20
wsuser.column_dimensions['B'].width = 12
wsuser.column_dimensions['C'].width = 12
wsuser.column_dimensions['D'].width = 20
wsuser.column_dimensions['E'].width = 20
wsuser.row_dimensions[1].height = 20
wsuser['A1'] = '时间'
wsuser['B1'] = '事件ID'
wsuser['C1'] = '操作'
wsuser['D1'] = '账号'
wsuser['E1'] = '操作账号'

# 样式
font = Font(size=10, name='宋体')
thin = Side(border_style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
# 对齐
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
title_font = Font(size=12, bold=True, name='宋体', color= 'ff0000')
for title_style in wslogin['A1:G1']:
    for title_cell in title_style:
        title_cell.font = title_font
        title_cell.border = border
        title_cell.alignment = alignment

for title_style in wsuser['A1:E1']:
    for title_cell in title_style:
        title_cell.font = title_font
        title_cell.border = border
        title_cell.alignment = alignment

EvtxPath = "Security.evtx"
i = 1
ii = 1
with Evtx.Evtx(EvtxPath) as log:
    size = len(list(log.records()))
    with alive_bar(size, title='分析进度') as bar:
        for record,sign in zip(log.records(),range(size)):
            xml = record.xml().replace(''' xmlns="http://schemas.microsoft.com/win/2004/08/events/event"''',"")
            tree = ElementTree.fromstring(xml)
            #事件ID
            EventID = tree.findall('System')[0].find('EventID').text
            if int(EventID) == 4624 or int(EventID) == 4625:
                #事件时间
                logtime = tree.findall('System')[0].find('TimeCreated').attrib['SystemTime'][0:19]
                logtime = datetime.strptime(logtime, "%Y-%m-%d %H:%M:%S") + timedelta(hours=8)
                for data in tree.findall('EventData')[0].findall('Data'):
                    if data.get("Name") == 'TargetUserName':
                        user = data.text
                    elif data.get("Name") == 'LogonType':
                        LogonType = data.text
                    elif data.get("Name") == 'IpAddress':
                        sourceip = data.text
                    elif data.get("Name") == 'ProcessName':
                        ProcessName = data.text
                wslogin.append([logtime,str(EventID),EventID.replace('4624','登陆成功').replace('4625','登录失败'),user,sourceip,int(LogonType),ProcessName])
                i += 1

            if int(EventID) == 4720 or int(EventID) == 4725 or int(EventID) == 4726:
                logtime = tree.findall('System')[0].find('TimeCreated').attrib['SystemTime'][0:19]
                logtime = datetime.strptime(logtime, "%Y-%m-%d %H:%M:%S") + timedelta(hours=8)
                for data in tree.findall('EventData')[0].findall('Data'):
                    if data.get("Name") == 'TargetUserName':
                        TargetUserName = data.text
                    elif data.get("Name") == 'SubjectUserName':
                        SubjectUserName = data.text
                wsuser.append([logtime,str(EventID),EventID.replace('4720','创建帐户').replace('4725','禁用帐户').replace('4726','删除账户'),TargetUserName,SubjectUserName])
                ii += 1
            bar()

        for row in wslogin['A2:G{}'.format(i)]:
            for cell in row:
                cell.font = font
                cell.border = border
                cell.alignment = alignment
        for row in wsuser['A2:E{}'.format(ii)]:
            for cell in row:
                cell.font = font
                cell.border = border
                cell.alignment = alignment
wslogin.auto_filter.add_sort_condition('G{}:A2'.format(i))
wslogin.auto_filter.add_sort_condition('F{}:A2'.format(ii))
wb.save(filename)
print(f'保存文件：{filename}')
time.sleep(3)